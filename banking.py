from openpyxl import load_workbook
from openpyxl import Workbook
import re

# Inserisci nome file da analizzare
# WORK_BOOK = "prova.xls"
WORK_BOOK = ["estrattogenn2k22-giu2k22.xlsx"]

# Inserisci nome file di destinazione
# DEST_FILE = "prova_destinazione.xlsx"
DEST_FILE = "estratto_fixed.xlsx"

# Modifica solo se sai cosa fai
FIXED_COSTS = {"(C/O)\s[a-zA-Z']*", "Causale:\s[\w0-9]*"}
MONTHS = {1:"Gennaio", 2:"Febbraio", 3:"Marzo", 4:"Aprile", 5:"Maggio", 6:"Giugno",
          7:"Luglio", 8:"Agosto", 9:"Settembre", 10: "Ottobre", 11:"Novembre", 12:"Dicembre"}
COLUMNS = {"Data" : "A","Uscite" : "C", "Entrate": "D", "Dettagli" : "G"}

def construct_workbook(): 
    output_wb = Workbook()
    output_wb.iso_dates = True
    output_wb_principal = output_wb.active
    output_wb_principal.title = "Main Page"
    output_wb.create_sheet("In-Out")
    output_wb["In-Out"].append(["Months", "Incomes", "Outcomes", "Delta"])
        
    for month in MONTHS:
        output_wb["In-Out"].append([MONTHS[month], 0, 0])
        output_wb.create_sheet(MONTHS[month])
        output_wb[f"{MONTHS[month]}"].append(list(COLUMNS.keys()))
    
    return output_wb

def load_input_sheet(work_book):
    input_wb = load_workbook(work_book)
    input_wb.iso_dates = True
    input_current_sheet = input_wb["estrattocontoitalia"]
    
    return input_current_sheet

def in_out_page(income, outcome, month, wb):
    wb = wb["In-Out"]
    income_sheet, outcome_sheet = int(wb.cell(month+1, 2).value), int(wb.cell(month+1, 3).value)
    wb.cell(month+1, 2, income+income_sheet)
    wb.cell(month+1, 3, outcome+outcome_sheet)
    wb.cell(month+1, 4, (income+income_sheet)+(outcome+outcome_sheet))
    
    return
    
def iter_rows(work_sheet):
    for row in work_sheet.iter_rows():
        yield [cell.value for cell in row]
        
        
        
        
        
if __name__ == "__main__":
    output_wb = construct_workbook()
        
    for work_book in WORK_BOOK:
        input_current_sheet = load_input_sheet(work_book)
        
        input_rows = list(iter_rows(input_current_sheet))
        input_rows.pop(0)
        
        for row in input_rows:
            current_month = int(row[0].month)            
            in_out_page((row[3] if row[3] is not None else 0), (row[2] if row[2] is not None else 0), current_month, output_wb)            
            current_output_ws = output_wb[f"{MONTHS[current_month]}"]
            
            if row[6] is not None:
                for cost in FIXED_COSTS:
                    match = re.search(cost, row[6])
                    if match is not None:
                        row[6] = row[6][match.start():match.end()]
            
            row = [str(row[0].strftime("%d/%m/%Y")), row[2], row[3], row[6]]
            current_output_ws.append(row)


output_wb.save(DEST_FILE)